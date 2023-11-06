import re
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def post_processing(results, savePath):
    '''
    This function makes xlsx files from Boxes and Texts.
    1. Loop Boxes and Texts according to document.
    2. In a document, loop boxes and texts according to page.
    3. SR number modification
    4. Consider cell swrap, thin
    5. Save
    '''

    col_title = ['Assembly Constituency number','Assembly Constituency name','Part Number','PDF Year','Main Town','Tehsil','District','Pin code','PDF Address', 'Voter ID number','Voter name','Husband/Father name', 'List of all matching T No', 'Page No.']
    pre_rows = 0 # considering multi tables in a page.
    wb = Workbook()
    ws = wb.active
    ws.title = "new table"
    for i in range(len(col_title)):
        ws.cell(pre_rows+1,i+1).value = col_title[i]
        ws.cell(pre_rows+1,i+1).font = Font(bold=True)    
    pre_rows += 1
    thin_border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')) 
    details = results["DETAILS"]
    for detail in details:
        
        ws.cell(row=pre_rows+1, column=1).value = results['assembly_number']
        ws.cell(row=pre_rows+1, column=2).value = results['assembly_name']
        ws.cell(row=pre_rows+1, column=3).value = results['part_number']
        ws.cell(row=pre_rows+1, column=4).value = results['year']
        ws.cell(row=pre_rows+1, column=5).value = results['main_town']
        ws.cell(row=pre_rows+1, column=6).value = results['tehsil']
        ws.cell(row=pre_rows+1, column=7).value = results['district']
        ws.cell(row=pre_rows+1, column=8).value = results['pin_code']
        ws.cell(row=pre_rows+1, column=9).value = results['address']
    

        ws.cell(row=pre_rows+1, column=10).value = detail['id']
        ws.cell(row=pre_rows+1, column=11).value = detail['name']
        ws.cell(row=pre_rows+1, column=12).value = detail['father_name']
        ws.cell(row=pre_rows+1, column=13).value = detail['house_no']
        ws.cell(row=pre_rows+1, column=14).value = detail['PageNumber']
        pre_rows += 1    

    # cell swrap, thin
    row_no = 1
    for i in ws.rows:
        for j in range(len(i)):
            ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
            ws.cell(row=row_no, column=j + 1).border = thin_border
        row_no = row_no + 1  
    column_width = [20, 20, 15, 12, 25, 20, 20, 20, 35, 25, 25, 25, 20, 8]

    for i in range(len(col_title)):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width[i]
    ws.sheet_view.zoomScale = 85

    wb.save(savePath)

    return None